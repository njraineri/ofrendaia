# --- Importaciones ---
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import database as db # Importa nuestro módulo database.py
import openpyxl
from io import BytesIO
import datetime
import os # Importar os para verificar existencia de archivo de DB

# --- Crear la instancia de la aplicación Flask ---
app = Flask(__name__)
app.secret_key = 'cambia_esta_clave_por_algo_seguro_y_unico' # ¡IMPORTANTE: Cambia esta clave!

# --- Filtro personalizado para formatear fechas ---
@app.template_filter('strftime')
def _jinja2_filter_datetime(date, fmt='%Y-%m-%d %H:%M:%S'):
    """Formatea un objeto datetime a una cadena usando strftime."""
    if isinstance(date, datetime.datetime):
        return date.strftime(fmt)
    return date # Retorna el valor sin modificar si no es un objeto datetime

# --- Context Processor (para hacer funciones/variables globales en Jinja2) ---
@app.context_processor
def inject_global_vars():
    """Hace disponibles variables o funciones globales en todas las plantillas."""
    try:
        if not os.path.exists(db.DATABASE_FILE):
             db.init_db()
        # Obtenemos los líderes aquí para que estén disponibles en plantillas como base.html
        # y _form_offering_modal.html. Manejamos posibles errores de DB.
        all_leaders_list = db.get_all_leaders()
    except Exception as e:
        # En un entorno de producción, querrías loggear esto de manera más robusta.
        print(f"ADVERTENCIA: No se pudieron cargar líderes en context_processor: {e}")
        all_leaders_list = [] # Retorna lista vacía si hay error con la DB


    return dict(
        get_all_leaders=lambda: all_leaders_list, # Retornamos una lambda que devuelve la lista obtenida o vacía
        now=datetime.datetime.utcnow # Función para obtener la hora actual (UTC)
    )

# --- Rutas Principales (Ventanas) ---

@app.route('/')
def index():
    """Página principal, redirige a la gestión de líderes."""
    return redirect(url_for('manage_leaders'))

@app.route('/leaders')
def manage_leaders():
    """Página para ver, buscar, agregar, editar y eliminar Líderes (ABM)."""
    search_term = request.args.get('search', '').strip()
    try:
        leaders = db.get_all_leaders(search_term)
    except Exception as e:
        flash(f"Error al cargar líderes: {e}", "danger")
        leaders = [] # Aseguramos que 'leaders' siempre sea una lista iterable
    # Pasamos todos los líderes al template para poder mostrarlos y quizás usarlos en modales de edición
    return render_template('leaders.html', leaders=leaders, search_term=search_term)


@app.route('/history')
def view_history():
    """Página para ver el histórico de Ofrendas, con filtros."""
    # Obtener parámetros de filtro de la URL
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    leader_id_filter = request.args.get('leader_id', 'all').strip() # 'all' por defecto

    # Validar y formatear fechas (simple)
    valid_start_date = start_date
    valid_end_date = end_date
    try:
        if start_date:
            datetime.datetime.strptime(start_date, '%Y-%m-%d')
    except ValueError:
        flash('Formato de fecha inicial inválido. Use %Y-%m-%d.', 'warning')
        valid_start_date = '' # Resetear si es inválido

    try:
        if end_date:
            datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        flash('Formato de fecha final inválido. Use %Y-%m-%d.', 'warning')
        valid_end_date = '' # Resetear si es inválido

    # Obtener historial filtrado y todos los líderes para el dropdown
    try:
        history = db.get_offerings_history(valid_start_date, valid_end_date, leader_id_filter)
        # all_leaders ya está disponible globalmente via context processor,
        # pero la pasamos aquí explícitamente también por claridad y si el context processor falla
        all_leaders = db.get_all_leaders()

        # --- Calcular el total de las ofrendas aquí en Python ---
        total_amount = 0.0
        for record in history:
             try:
                 # Corregido: Usar acceso por corchetes [] en lugar de .get()
                 # Asegurarse de que 'amount' sea un número antes de sumarlo
                 total_amount += float(record['amount'])
             except (ValueError, TypeError):
                 # Si hay un problema con el valor, lo ignoramos o lo tratamos como 0
                 print(f"ADVERTENCIA: Valor de ofrenda no numérico encontrado: {record.get('amount')}")
                 pass # Ignorar valores no numéricos


    except Exception as e:
         flash(f"Error al cargar el historial o los líderes: {e}", "danger")
         history = []
         all_leaders = [] # Aseguramos que 'all_leaders' siempre sea una lista iterable
         total_amount = 0.0 # Resetear total si hay error

    # Pasamos el historial, los líderes y el total calculado a la plantilla
    return render_template('history.html',
                           history=history,
                           leaders=all_leaders, # Pasamos todos los líderes
                           start_date=start_date,
                           end_date=end_date,
                           selected_leader_id=leader_id_filter,
                           total_amount=total_amount) # >>> PASAMOS EL TOTAL <<<

# --- Rutas para Acciones ABM Líderes ---

@app.route('/leaders/add', methods=['POST'])
def add_leader():
    """Procesa el formulario para agregar un nuevo líder."""
    name = request.form.get('leader_name', '').strip()
    direccion = request.form.get('leader_direccion', '').strip()
    dia = request.form.get('leader_dia', '').strip()
    hora = request.form.get('leader_hora', '').strip()

    if not name:
        flash('El nombre del líder no puede estar vacío.', 'warning')
    else:
        try:
            if db.add_leader(name, direccion, dia, hora):
                flash(f'Líder "{name}" agregado exitosamente.', 'success')
            else:
                flash(f'Error: El líder "{name}" ya existe.', 'danger')
        except Exception as e:
            flash(f"Error inesperado al agregar líder: {e}", "danger")

    return redirect(url_for('manage_leaders'))

@app.route('/leaders/edit/<int:leader_id>', methods=['POST'])
def edit_leader(leader_id):
    """Procesa el formulario para editar un líder existente."""
    new_name = request.form.get('leader_name', '').strip()
    new_direccion = request.form.get('leader_direccion', '').strip()
    new_dia = request.form.get('leader_dia', '').strip()
    new_hora = request.form.get('leader_hora', '').strip()


    if not new_name:
        flash('El nombre del líder no puede estar vacío.', 'warning')
        return redirect(url_for('manage_leaders'))

    try:
        original_leader = db.get_leader_by_id(leader_id)
        if not original_leader:
            flash('Error: Líder no encontrado para editar.', 'danger')
        else:
            if db.update_leader(leader_id, new_name, new_direccion, new_dia, new_hora):
                 # Verificar si hubo cambios para mostrar un mensaje más informativo
                 if (new_name != original_leader.get('name') or
                     new_direccion != original_leader.get('direccion') or
                     new_dia != original_leader.get('dia') or
                     new_hora != original_leader.get('hora')):
                     flash(f'Líder "{original_leader.get("name", "ID " + str(leader_id))}" actualizado exitosamente.', 'success')
                 else:
                      # No se realizaron cambios
                      pass # No mostramos mensaje si no hay cambios
            else:
                # update_leader podría devolver False si el nuevo nombre ya existe
                flash(f'Error: Ya existe otro líder con el nombre "{new_name}".', 'danger')
    except Exception as e:
        flash(f"Error inesperado al editar líder: {e}", "danger")

    return redirect(url_for('manage_leaders'))


@app.route('/leaders/delete/<int:leader_id>', methods=['POST'])
def delete_leader(leader_id):
    """Procesa la solicitud para eliminar un líder."""
    try:
        leader = db.get_leader_by_id(leader_id) # Obtener antes de borrar para el mensaje
        if leader:
            if db.delete_leader(leader_id):
                flash(f'Líder "{leader.get("name", "ID " + str(leader_id))}" y sus ofrendas asociadas han sido eliminados.', 'success')
            else:
                flash(f'Error al intentar eliminar al líder "{leader.get("name", "ID " + str(leader_id))}". Consulte los logs.', 'danger')
        else:
            flash('Error: Líder no encontrado para eliminar.', 'danger')
    except Exception as e:
        flash(f"Error inesperado al eliminar líder: {e}", "danger")

    return redirect(url_for('manage_leaders'))

# --- Ruta para Acción Registrar Ofrenda ---

@app.route('/offerings/add', methods=['POST'])
def add_offering():
    """Procesa el formulario del pop-up para registrar una nueva ofrenda."""
    amount_str = request.form.get('offering_amount', '0').replace(',', '.').strip()
    date = request.form.get('offering_date', '').strip()
    leader_id = request.form.get('offering_leader', '').strip()

    error = False
    if not date:
        flash('La fecha es obligatoria.', 'warning')
        error = True
    if not leader_id or leader_id == 'all': # Validar que se seleccione un líder específico
        flash('Debe seleccionar un líder válido.', 'warning')
        error = True

    amount = 0.0
    try:
        amount = float(amount_str)
        if amount <= 0:
             flash('El monto debe ser un número positivo.', 'warning')
             error = True
    except ValueError:
        flash('Monto inválido. Ingrese un número válido (ej: 150.50).', 'danger')
        error = True

    if not error:
        try:
            # Convertir leader_id a entero antes de pasarlo a la DB
            leader_id_int = int(leader_id)
            if db.add_offering(amount, date, leader_id_int):
                flash('Ofrenda registrada exitosamente.', 'success')
            else:
                flash('Error al intentar registrar la ofrenda en la base de datos.', 'danger')
        except ValueError:
             flash('ID de líder inválido.', 'danger') # Si el ID no es un número válido
        except Exception as e:
            flash(f'Error inesperado al registrar la ofrenda: {e}', 'danger')

    # Redirigir a la página de historial para ver el resultado
    # Mantenemos los filtros actuales si es posible (asumiendo que se pasan en el form)
    start_date = request.form.get('start_date', '').strip()
    end_date = request.form.get('end_date', '').strip()
    leader_id_filter = request.form.get('leader_id_filter', 'all').strip()

    return redirect(url_for('view_history',
                            start_date=start_date,
                            end_date=end_date,
                            leader_id=leader_id_filter))


# --- Nuevas Rutas para Acciones de Ofrendas ---

@app.route('/offerings/edit/<int:offering_id>', methods=['POST'])
def edit_offering(offering_id):
    """Procesa el formulario para editar una ofrenda existente."""
    amount_str = request.form.get('offering_amount', '0').replace(',', '.').strip()
    date = request.form.get('offering_date', '').strip()
    leader_id = request.form.get('offering_leader', '').strip()

    error = False
    if not date:
        flash('La fecha es obligatoria.', 'warning')
        error = True
    if not leader_id or leader_id == 'all':
        flash('Debe seleccionar un líder válido.', 'warning')
        error = True

    amount = 0.0
    try:
        amount = float(amount_str)
        if amount <= 0:
             flash('El monto debe ser un número positivo.', 'warning')
             error = True
    except ValueError:
        flash('Monto inválido. Ingrese un número válido (ej: 150.50).', 'danger')
        error = True

    if not error:
        try:
            leader_id_int = int(leader_id)
            if db.update_offering(offering_id, amount, date, leader_id_int):
                 flash('Ofrenda actualizada exitosamente.', 'success')
            else:
                 flash('Error al intentar actualizar la ofrenda.', 'danger')
        except ValueError:
             flash('ID de líder o ID de ofrenda inválido.', 'danger')
        except Exception as e:
            flash(f'Error inesperado al actualizar la ofrenda: {e}', 'danger')

    # Redirigir de vuelta a la página de historial
    # Mantenemos los filtros actuales si es posible
    start_date = request.form.get('start_date', '').strip()
    end_date = request.form.get('end_date', '').strip()
    leader_id_filter = request.form.get('leader_id_filter', 'all').strip()

    return redirect(url_for('view_history',
                            start_date=start_date,
                            end_date=end_date,
                            leader_id=leader_id_filter))


@app.route('/offerings/delete/<int:offering_id>', methods=['POST'])
def delete_offering(offering_id):
    """Procesa la solicitud para eliminar una ofrenda."""
    try:
        if db.delete_offering(offering_id):
            flash('Ofrenda eliminada exitosamente.', 'success')
        else:
            flash('Error al intentar eliminar la ofrenda.', 'danger')
    except Exception as e:
        flash(f"Error inesperado al eliminar ofrenda: {e}", "danger")

    # Redirigir de vuelta a la página de historial
    # Mantenemos los filtros actuales si es posible (asumiendo que se pasan en el form)
    start_date = request.form.get('start_date', '').strip()
    end_date = request.form.get('end_date', '').strip()
    leader_id_filter = request.form.get('leader_id_filter', 'all').strip()

    return redirect(url_for('view_history',
                            start_date=start_date,
                            end_date=end_date,
                            leader_id=leader_id_filter))


# --- Ruta para Exportar a Excel ---

@app.route('/export')
def export_history():
    """Genera y descarga un archivo Excel con el histórico filtrado."""
    # Reutilizar los mismos filtros que la vista de historial
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    leader_id_filter = request.args.get('leader_id', 'all').strip()

    # Obtener los datos filtrados
    try:
        history = db.get_offerings_history(start_date, end_date, leader_id_filter)
    except Exception as e:
        flash(f"Error al obtener datos para exportar: {e}", "danger")
        return redirect(url_for('view_history')) # Volver al historial si hay error

    # Crear el archivo Excel en memoria
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Histórico Ofrendas"

    # Escribir cabeceras
    headers = ["Fecha", "Líder", "Monto"]
    sheet.append(headers)
    # Aplicar negrita a cabeceras (opcional)
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)


    # Escribir datos
    total_amount = 0.0
    for record in history:
        try:
            # Corregido: Usar acceso por corchetes [] en lugar de .get()
            amount_float = float(record['amount'])
            total_amount += amount_float
        except (ValueError, TypeError):
            amount_float = 0.0 # O dejar como texto original: record['amount']
            print(f"ADVERTENCIA: Valor de ofrenda no numérico encontrado durante exportación: {record.get('amount')}")


        sheet.append([
            record['date'],
            record['leader_name'],
            amount_float
        ])

    # Añadir fila de total (opcional)
    if history:
        total_row_idx = len(history) + 2 # Índice de la fila después de la última data
        sheet.cell(row=total_row_idx, column=2, value="Total Filtrado:").font = openpyxl.styles.Font(bold=True)
        total_cell = sheet.cell(row=total_row_idx, column=3, value=total_amount)
        total_cell.font = openpyxl.styles.Font(bold=True)
        total_cell.number_format = '$ #,##0.00' # Formato de moneda para el total


    # Ajustar ancho de columnas
    sheet.column_dimensions['A'].width = 12 # Fecha
    sheet.column_dimensions['B'].width = 30 # Líder (puede ser largo)
    sheet.column_dimensions['C'].width = 15 # Monto

    # Aplicar formato de moneda a la columna C (Monto)
    currency_format = '$ #,##0.00' # Ajusta según tu moneda
    for row_idx in range(2, len(history) + 2): # Desde la fila 2 hasta la última de datos
        sheet.cell(row=row_idx, column=3).number_format = currency_format

    # Guardar en un buffer de BytesIO
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0) # Regresar al inicio del buffer

    # Generar nombre de archivo dinámico
    filename_parts = ["historico_ofrendas"]
    if start_date: filename_parts.append(f"desde_{start_date}")
    if end_date: filename_parts.append(f"hasta_{end_date}")
    leader_name = "todos"
    if leader_id_filter != 'all':
        try:
            leader = db.get_leader_by_id(int(leader_id_filter))
            if leader: leader_name = leader['name'].replace(" ", "_").lower() # Nombre amigable
            filename_parts.append(f"lider_{leader_name}")
        except (ValueError, TypeError):
             filename_parts.append(f"lider_id_{leader_id_filter}") # Si el ID no es válido

    filename = "_".join(filename_parts) + ".xlsx"

    # Enviar el archivo al navegador para descarga
    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# --- Bloque para Ejecutar la Aplicación ---
if __name__ == '__main__':
    print("--- Sistema de Gestión de Ofrendas ---")
    # La inicialización de la DB se maneja al importar database.py
    print("Base de datos lista.")
    print("Iniciando servidor de desarrollo Flask...")
    print(f"Accede a la aplicación en tu navegador web:")
    print(f"  - Localmente: http://127.0.0.1:5000")
    print(f"  - En red local: http://<IP_DE_TU_MAQUINA>:5000")
    print("Presiona CTRL+C para detener el servidor.")
    # Considera añadir host='0.0.0.0' si necesitas acceder desde otros dispositivos
    app.run(debug=True, host='0.0.0.0', port=5000)