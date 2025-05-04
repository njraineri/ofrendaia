# app.py COMPLETO Y CORRECTO (con filtros mes/año y TODAS las rutas)

from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import database as db
import openpyxl
from io import BytesIO
import datetime
import os
import calendar

app = Flask(__name__)
app.secret_key = 'cambia_esta_clave_por_algo_seguro_y_unico'

@app.template_filter('strftime')
def _jinja2_filter_datetime(date, fmt='%Y-%m-%d %H:%M:%S'):
    if isinstance(date, datetime.datetime):
        return date.strftime(fmt)
    return date

@app.context_processor
def inject_global_vars():
    try:
        if not os.path.exists(db.DATABASE_FILE):
            db.init_db()
        all_leaders_list = db.get_all_leaders()
    except Exception as e:
        print(f"ADVERTENCIA: No se pudieron cargar líderes en context_processor: {e}")
        all_leaders_list = []
    return dict(get_all_leaders=lambda: all_leaders_list, now=datetime.datetime.utcnow)

@app.route('/')
def index():
    return redirect(url_for('manage_leaders'))

@app.route('/leaders')
def manage_leaders():
    search_term = request.args.get('search', '').strip()
    try:
        leaders = db.get_all_leaders(search_term)
    except Exception as e:
        flash(f"Error al cargar líderes: {e}", "danger")
        leaders = []
    return render_template('leaders.html', leaders=leaders, search_term=search_term)

@app.route('/leaders/add', methods=['POST'])
def add_leader():
    name = request.form.get('leader_name', '').strip()
    direccion = request.form.get('leader_direccion', '').strip()
    dia = request.form.get('leader_dia', '').strip()
    hora = request.form.get('leader_hora', '').strip()

    if not name:
        flash('El nombre del líder no puede estar vacío.', 'warning')
    else:
        try:
            if db.add_leader(name, direccion, dia, hora):
                flash(f'Líder "{name}" agregado exitosamente.', 'success')
            else:
                flash(f'Error: El líder "{name}" ya existe.', 'danger')
        except Exception as e:
            flash(f"Error inesperado al agregar líder: {e}", "danger")

    return redirect(url_for('manage_leaders'))

@app.route('/leaders/edit/<int:leader_id>', methods=['POST'])
def edit_leader(leader_id):
    new_name = request.form.get('leader_name', '').strip()
    new_direccion = request.form.get('leader_direccion', '').strip()
    new_dia = request.form.get('leader_dia', '').strip()
    new_hora = request.form.get('leader_hora', '').strip()

    if not new_name:
        flash('El nombre del líder no puede estar vacío.', 'warning')
        return redirect(url_for('manage_leaders'))

    try:
        original_leader = db.get_leader_by_id(leader_id)
        if not original_leader:
            flash('Error: Líder no encontrado para editar.', 'danger')
        else:
            if db.update_leader(leader_id, new_name, new_direccion, new_dia, new_hora):
                flash(f'Líder "{original_leader.get("name", "ID " + str(leader_id))}" actualizado.', 'success')
            else:
                flash(f'Error: Ya existe otro líder con el nombre "{new_name}".', 'danger')
    except Exception as e:
        flash(f"Error inesperado al editar líder: {e}", "danger")

    return redirect(url_for('manage_leaders'))

@app.route('/leaders/delete/<int:leader_id>', methods=['POST'])
def delete_leader(leader_id):
    try:
        leader = db.get_leader_by_id(leader_id)
        if leader:
            if db.delete_leader(leader_id):
                flash(f'Líder "{leader.get("name", "ID " + str(leader_id))}" eliminado.', 'success')
            else:
                flash(f'Error al eliminar líder.', 'danger')
        else:
            flash('Error: Líder no encontrado.', 'danger')
    except Exception as e:
        flash(f"Error inesperado al eliminar líder: {e}", "danger")

    return redirect(url_for('manage_leaders'))

@app.route('/history')
def view_history():
    start_month = request.args.get('start_month', type=int)
    end_month = request.args.get('end_month', type=int)
    year = request.args.get('year', type=int)
    leader_id_filter = request.args.get('leader_id', 'all').strip()

    today = datetime.datetime.today()
    if not start_month:
        start_month = today.month
    if not end_month:
        end_month = today.month
    if not year:
        year = today.year

    start_date = datetime.date(year, start_month, 1)
    last_day = calendar.monthrange(year, end_month)[1]
    end_date = datetime.date(year, end_month, last_day)

    try:
        history = db.get_offerings_history(start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'), leader_id_filter)
        all_leaders = db.get_all_leaders()
        total_amount = sum(float(r['amount']) for r in history if r['amount'])
    except Exception as e:
        flash(f"Error al cargar historial o líderes: {e}", "danger")
        history, all_leaders, total_amount = [], [], 0.0

    return render_template('history.html', history=history, leaders=all_leaders,
                           start_month=start_month, end_month=end_month, year=year,
                           selected_leader_id=leader_id_filter, total_amount=total_amount)

@app.route('/offerings/add', methods=['POST'])
def add_offering():
    amount_str = request.form.get('offering_amount', '0').replace(',', '.').strip()
    date = request.form.get('offering_date', '').strip()
    leader_id = request.form.get('offering_leader', '').strip()

    error = False
    if not date:
        flash('La fecha es obligatoria.', 'warning')
        error = True
    if not leader_id or leader_id == 'all':
        flash('Debe seleccionar un líder válido.', 'warning')
        error = True

    try:
        amount = float(amount_str)
        if amount <= 0:
            flash('El monto debe ser positivo.', 'warning')
            error = True
    except ValueError:
        flash('Monto inválido.', 'danger')
        error = True

    if not error:
        try:
            leader_id_int = int(leader_id)
            if db.add_offering(amount, date, leader_id_int):
                flash('Ofrenda registrada.', 'success')
            else:
                flash('Error al registrar ofrenda.', 'danger')
        except Exception as e:
            flash(f'Error inesperado al registrar: {e}', 'danger')

    return redirect(url_for('view_history'))

@app.route('/offerings/edit/<int:offering_id>', methods=['POST'])
def edit_offering(offering_id):
    amount_str = request.form.get('offering_amount', '0').replace(',', '.').strip()
    date = request.form.get('offering_date', '').strip()
    leader_id = request.form.get('offering_leader', '').strip()

    error = False
    if not date:
        flash('La fecha es obligatoria.', 'warning')
        error = True
    if not leader_id or leader_id == 'all':
        flash('Debe seleccionar un líder válido.', 'warning')
        error = True

    try:
        amount = float(amount_str)
        if amount <= 0:
            flash('El monto debe ser positivo.', 'warning')
            error = True
    except ValueError:
        flash('Monto inválido.', 'danger')
        error = True

    if not error:
        try:
            leader_id_int = int(leader_id)
            if db.update_offering(offering_id, amount, date, leader_id_int):
                flash('Ofrenda actualizada.', 'success')
            else:
                flash('Error al actualizar ofrenda.', 'danger')
        except Exception as e:
            flash(f'Error inesperado al actualizar: {e}', 'danger')

    return redirect(url_for('view_history'))

@app.route('/offerings/delete/<int:offering_id>', methods=['POST'])
def delete_offering(offering_id):
    try:
        if db.delete_offering(offering_id):
            flash('Ofrenda eliminada.', 'success')
        else:
            flash('Error al eliminar ofrenda.', 'danger')
    except Exception as e:
        flash(f"Error inesperado al eliminar: {e}", "danger")

    return redirect(url_for('view_history'))

@app.route('/export')
def export_history():
    start_month = request.args.get('start_month', type=int)
    end_month = request.args.get('end_month', type=int)
    year = request.args.get('year', type=int)
    leader_id_filter = request.args.get('leader_id', 'all').strip()

    today = datetime.datetime.today()
    if not start_month:
        start_month = today.month
    if not end_month:
        end_month = today.month
    if not year:
        year = today.year

    start_date = datetime.date(year, start_month, 1)
    last_day = calendar.monthrange(year, end_month)[1]
    end_date = datetime.date(year, end_month, last_day)

    try:
        history = db.get_offerings_history(start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'), leader_id_filter)
    except Exception as e:
        flash(f"Error al obtener datos para exportar: {e}", "danger")
        return redirect(url_for('view_history'))

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Histórico Ofrendas"

    headers = ["Fecha", "Líder", "Monto"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    total_amount = 0.0
    for record in history:
        try:
            amount_float = float(record['amount'])
            total_amount += amount_float
        except (ValueError, TypeError):
            amount_float = 0.0
        sheet.append([record['date'], record['leader_name'], amount_float])

    if history:
        total_row_idx = len(history) + 2
        sheet.cell(row=total_row_idx, column=2, value="Total Filtrado:").font = openpyxl.styles.Font(bold=True)
        total_cell = sheet.cell(row=total_row_idx, column=3, value=total_amount)
        total_cell.font = openpyxl.styles.Font(bold=True)
        total_cell.number_format = '$ #,##0.00'

    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 15

    currency_format = '$ #,##0.00'
    for row_idx in range(2, len(history) + 2):
        sheet.cell(row=row_idx, column=3).number_format = currency_format

    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    filename = f"ofrendas_{year}_{start_month}_a_{end_month}.xlsx"

    return send_file(excel_buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)